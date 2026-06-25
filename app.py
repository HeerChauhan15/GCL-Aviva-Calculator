import streamlit as st
import pandas as pd
import numpy as np
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Insurance Premium Calculator",
    page_icon="💰",
    layout="wide"
)

st.title("💰 Aviva GCL Insurance Premium Calculator")
st.markdown("Select plan details below")

st.warning("⚠️ It is mandatory to select **Life Type** (Single Life / Joint Life) and **Loan Type** (Home Loan / LAP) before proceeding.")

FILE_MAP = {
    ("Single Life", "Home Loan"): "Aviva Single HomeLoan.xlsx",
    ("Single Life", "LAP"):       "Aviva Single Lap.xlsx",
    ("Joint Life",  "Home Loan"): "Aviva Joint Homeloan.xlsx",
    ("Joint Life",  "LAP"):       "Aviva Joint Lap.xlsx",
}

def load_rate_table(life_type, loan_type):
    fname = FILE_MAP[(life_type, loan_type)]
    if not os.path.exists(fname):
        raise FileNotFoundError(
            f"File not found: '{fname}' — Please make sure this file is in the GitHub repo."
        )

    raw = pd.read_excel(fname, sheet_name="Sheet1", header=None)

    header_row = None
    for i, row in raw.iterrows():
        for val in row.values:
            if isinstance(val, str) and "AGE" in val.upper():
                header_row = i
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError("Could not find AGE/TERM header row in the file.")

    df = pd.read_excel(fname, sheet_name="Sheet1", header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    age_col = df.columns[0]
    df = df.dropna(subset=[age_col])
    df[age_col] = pd.to_numeric(df[age_col], errors='coerce')
    df = df.dropna(subset=[age_col])
    df[age_col] = df[age_col].astype(int)
    df = df.set_index(age_col)

    tenure_map = {}
    for col in df.columns:
        try:
            tenure_map[int(float(col))] = col
        except Exception:
            pass

    return df, tenure_map


def get_rate(df, tenure_map, age, tenure):
    if age not in df.index:
        raise ValueError(f"Age {age} not found in rate table.")
    if tenure not in tenure_map:
        raise ValueError(f"Tenure {tenure} yrs not found in rate table.")
    return float(df.loc[age, tenure_map[tenure]])


def find_column(df, target):
    """Exact (case/space-insensitive) match — used for Single Life."""
    target_norm = target.strip().lower().replace(" ", "")
    for col in df.columns:
        col_norm = str(col).strip().lower().replace(" ", "")
        if col_norm == target_norm:
            return col
    return None


def _normalize(s):
    return re.sub(r'[\s_\-]+', '', str(s).lower())


def detect_person(norm):
    """Decide whether a (normalized) column name belongs to Main Borrower or Co Borrower."""
    if 'mainborrower' in norm or norm.startswith('mb') or 'borrower1' in norm:
        return 'main'
    if 'coborrower' in norm or norm.startswith('cb') or 'borrower2' in norm:
        return 'co'
    # fallback: trailing 1 / 2 like Name1 / Name2, Age1 / Age2
    if norm.endswith('1'):
        return 'main'
    if norm.endswith('2'):
        return 'co'
    return None


def detect_field(norm):
    if 'name' in norm:
        return 'name'
    if 'tenure' in norm:
        return 'tenure'
    if 'age' in norm:
        return 'age'
    return None


def map_joint_columns(df):
    """
    Flexibly detect Main Borrower / Co Borrower Name/Age/Tenure columns
    regardless of exact header wording (Main Borrower Name, MB Name, Name1,
    Borrower 1 Name, etc.)

    Note: Co Borrower Tenure is intentionally NOT looked for — the loan tenure
    is shared, so Co Borrower's premium always uses Main Borrower's Tenure.
    """
    mapping = {}  # (person, field) -> actual column name
    for col in df.columns:
        norm = _normalize(col)
        field = detect_field(norm)
        if field is None:
            continue
        person = detect_person(norm)
        if person is None:
            continue
        if person == 'co' and field == 'tenure':
            # Co Borrower tenure is ignored; Main Borrower tenure is used for both.
            continue
        key = (person, field)
        if key not in mapping:
            mapping[key] = col
    return mapping


def find_sum_assured_column(df):
    """Flexibly detect an optional Sum Assured / Sum Insured / Loan Amount column."""
    for col in df.columns:
        norm = _normalize(col)
        if 'sumassured' in norm or 'suminsured' in norm or 'loanamount' in norm:
            return col
    return None


# ============================================
# DROPDOWNS
# ============================================

col1, col2 = st.columns(2)
with col1:
    life_type = st.selectbox("Select Life Type", ["Single Life", "Joint Life"])
with col2:
    loan_type = st.selectbox("Select Loan Type", ["Home Loan", "LAP"])

# ============================================
# SUM ASSURED (mandatory) — rates in the backend files are per ₹1,00,000
# ============================================
if loan_type == "Home Loan":
    sa_min, sa_max = 100000, 6000000
else:
    sa_min, sa_max = 100000, 4000000

sum_assured = st.number_input(
    "Select Sum Assured (₹)",
    min_value=sa_min,
    max_value=sa_max,
    value=sa_min,
    step=100000,
    help=f"For {loan_type}, Sum Assured must be between ₹{sa_min:,} and ₹{sa_max:,}."
)

st.divider()

# ============================================
# MANUAL SECTION
# ============================================

st.subheader("🔢 Manual Rate Lookup")

if loan_type == "Home Loan":
    min_tenure, max_tenure = 5, 25
else:
    min_tenure, max_tenure = 2, 10

if life_type == "Single Life":
    col3, col4 = st.columns(2)
    with col3:
        age = st.number_input("Enter Age", min_value=18, max_value=65, value=30, step=1)
    with col4:
        tenure = st.number_input(
            "Enter Tenure",
            min_value=min_tenure,
            max_value=max_tenure,
            value=min_tenure,
            step=1
        )
        st.caption("📅 Tenure is in Years")

    if st.button("Get Rate", type="primary"):
        try:
            df_rates, tenure_map = load_rate_table(life_type, loan_type)
            rate = get_rate(df_rates, tenure_map, age, tenure)
            premium = rate * (sum_assured / 100000)
            st.success(f"✅ {life_type} | {loan_type} | Age {age} | Tenure {tenure} yrs | Sum Assured ₹{sum_assured:,}")
            col_a, col_b = st.columns(2)
            with col_a:
                st.metric("Rate (per ₹1,00,000)", f"₹ {rate:,.2f}")
            with col_b:
                st.metric("Premium (for selected Sum Assured)", f"₹ {premium:,.2f}")
        except Exception as e:
            st.error(f"Error: {e}")

else:
    st.markdown("**Main Borrower**")
    mcol1, mcol2 = st.columns(2)
    with mcol1:
        main_age = st.number_input("Age", min_value=18, max_value=65, value=30, step=1, key="main_age_manual")
    with mcol2:
        main_tenure = st.number_input(
            "Tenure", min_value=min_tenure, max_value=max_tenure,
            value=min_tenure, step=1, key="main_tenure_manual"
        )
    st.caption("📅 Tenure is in Years")

    st.markdown("**Co Borrower**")
    co_age = st.number_input("Age", min_value=18, max_value=65, value=30, step=1, key="co_age_manual")
    st.caption("📅 Co Borrower Tenure is the same as Main Borrower Tenure (loan tenure is shared)")

    if st.button("Get Rate", type="primary"):
        try:
            df_rates, tenure_map = load_rate_table(life_type, loan_type)

            rate_main = get_rate(df_rates, tenure_map, main_age, main_tenure)
            premium_main = rate_main * (sum_assured / 100000)

            # Co Borrower uses Main Borrower's tenure (shared loan tenure)
            rate_co = get_rate(df_rates, tenure_map, co_age, main_tenure)
            premium_co = rate_co * (sum_assured / 100000)

            total_premium = premium_main + premium_co

            st.success(
                f"✅ {life_type} | {loan_type} | Sum Assured ₹{sum_assured:,} | "
                f"Main Borrower: Age {main_age}, Tenure {main_tenure} yrs | "
                f"Co Borrower: Age {co_age}, Tenure {main_tenure} yrs (same as Main Borrower)"
            )

            col_a, col_b, col_c = st.columns(3)
            with col_a:
                st.metric("Main Borrower Premium", f"₹ {premium_main:,.2f}")
            with col_b:
                st.metric("Co Borrower Premium", f"₹ {premium_co:,.2f}")
            with col_c:
                st.metric("Total Premium", f"₹ {total_premium:,.2f}")
        except Exception as e:
            st.error(f"Error: {e}")

st.divider()

# ============================================
# EXCEL UPLOAD SECTION
# ============================================

st.subheader("📂 Upload Member Data for Bulk Rate Lookup")

if life_type == "Single Life":
    st.markdown(
        "Your Excel must have at least: **Name**, **Age**, **Tenure** (in years). "
        "You may also include a **Sum Assured** column — if not provided, the Sum Assured "
        "selected above will be used for all entries."
    )
else:
    st.markdown(
        "Your Excel must have at least: **Main Borrower** (Name, Age, Tenure in years) and "
        "**Co Borrower** (Name, Age). "
        "Co Borrower's Tenure is **not required** — it is always assumed to be the same as "
        "Main Borrower's Tenure, since the loan tenure is shared between borrowers. "
        "If your file has a Co Borrower Tenure column, it will simply be ignored. "
        "You may also include a **Sum Assured** column — if not provided, the Sum Assured "
        "selected above will be used for all entries."
    )

st.warning("⚠️ Please make sure you have selected **Life Type** and **Loan Type** above before uploading your Excel file.")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file)
        df.columns = [str(c).strip() for c in df.columns]

        st.subheader("Uploaded Data Preview")
        st.dataframe(df.head())

        if loan_type == "Home Loan":
            min_t, max_t = 5, 25
        else:
            min_t, max_t = 2, 10

        df_rates, tenure_map = load_rate_table(life_type, loan_type)

        # ============================================
        # SINGLE LIFE
        # ============================================
        if life_type == "Single Life":
            name_col = find_column(df, "Name")
            age_col = find_column(df, "Age")
            tenure_col = find_column(df, "Tenure")
            sa_col = find_sum_assured_column(df)

            if not name_col or not age_col or not tenure_col:
                raise ValueError("Excel must contain mandatory columns: Name, Age, Tenure")

            df[age_col] = pd.to_numeric(df[age_col], errors='coerce')
            df[tenure_col] = pd.to_numeric(df[tenure_col], errors='coerce')

            if df[tenure_col].dropna().median() > 30:
                st.info("ℹ️ Tenure values look like months — auto-converting to years.")
                df[tenure_col] = (df[tenure_col] / 12).round(0).astype('Int64')
            else:
                df[tenure_col] = df[tenure_col].round(0).astype('Int64')

            df[age_col] = df[age_col].round(0).astype('Int64')
            df[tenure_col] = df[tenure_col].clip(lower=min_t, upper=max_t)

            if sa_col:
                st.info(f"ℹ️ Found '{sa_col}' column — using per-row Sum Assured (capped to ₹{sa_min:,}–₹{sa_max:,}).")
                df[sa_col] = pd.to_numeric(df[sa_col], errors='coerce').fillna(sum_assured)
                df[sa_col] = df[sa_col].clip(lower=sa_min, upper=sa_max)
                sa_series = df[sa_col]
            else:
                sa_series = pd.Series([sum_assured] * len(df), index=df.index)

            premiums = []
            statuses = []
            for idx, row in df.iterrows():
                try:
                    r_age = int(row[age_col])
                    r_tenure = int(row[tenure_col])
                    r = get_rate(df_rates, tenure_map, r_age, r_tenure)
                    row_sa = float(sa_series.loc[idx])
                    premium = round(r * (row_sa / 100000), 2)
                    premiums.append(premium)
                    statuses.append("✅")
                except Exception as e:
                    premiums.append(None)
                    statuses.append(f"❌ {e}")

            df["Premium"] = premiums
            df["Status"] = statuses

            core_cols = [name_col, age_col, tenure_col, "Premium"]
            extra_cols = [c for c in df.columns if c not in core_cols]
            df = df[core_cols + extra_cols]

            total_premium = pd.to_numeric(pd.Series(premiums), errors='coerce').sum()
            st.metric("💰 Grand Total Premium", f"₹ {total_premium:,.2f}")

            st.subheader("Rate Lookup Output")
            st.dataframe(df, use_container_width=True)

            total_row = {c: "" for c in df.columns}
            total_row[name_col] = "TOTAL PREMIUM"
            total_row["Premium"] = round(total_premium, 2)
            df_out = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

            output_file = "Rate_Output.xlsx"
            df_out.to_excel(output_file, index=False)

            with open(output_file, "rb") as file:
                st.download_button(
                    label="⬇ Download Output Excel",
                    data=file,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        # ============================================
        # JOINT LIFE
        # ============================================
        else:
            mapping = map_joint_columns(df)
            sa_col = find_sum_assured_column(df)

            # Note: Co Borrower Tenure is intentionally not required — the loan
            # tenure is shared, so Main Borrower's Tenure is used for both.
            required_keys = [
                ('main', 'name'), ('main', 'age'), ('main', 'tenure'),
                ('co', 'name'), ('co', 'age'),
            ]
            missing = [f"{p.capitalize()} Borrower {f.capitalize()}" for (p, f) in required_keys if (p, f) not in mapping]

            if missing:
                raise ValueError(
                    "Could not detect these mandatory columns: " + ", ".join(missing) +
                    ". Please make sure your Excel has Main Borrower & Co Borrower "
                    "Name/Age columns and a Main Borrower Tenure column (any reasonable naming "
                    "works, e.g. 'Main Borrower Age', 'MB Age', 'Age1'). Co Borrower Tenure is "
                    "not needed — Main Borrower's Tenure is used for both borrowers."
                )

            main_name_col = mapping[('main', 'name')]
            main_age_col = mapping[('main', 'age')]
            main_tenure_col = mapping[('main', 'tenure')]
            co_name_col = mapping[('co', 'name')]
            co_age_col = mapping[('co', 'age')]

            for c in [main_age_col, main_tenure_col, co_age_col]:
                df[c] = pd.to_numeric(df[c], errors='coerce')

            if df[main_tenure_col].dropna().median() > 30:
                st.info("ℹ️ Main Borrower Tenure values look like months — auto-converting to years.")
                df[main_tenure_col] = (df[main_tenure_col] / 12).round(0).astype('Int64')
            else:
                df[main_tenure_col] = df[main_tenure_col].round(0).astype('Int64')

            df[main_age_col] = df[main_age_col].round(0).astype('Int64')
            df[co_age_col] = df[co_age_col].round(0).astype('Int64')
            df[main_tenure_col] = df[main_tenure_col].clip(lower=min_t, upper=max_t)

            st.info("ℹ️ Co Borrower Tenure is assumed to be the same as Main Borrower Tenure for all rows.")

            if sa_col:
                st.info(f"ℹ️ Found '{sa_col}' column — using per-row Sum Assured (capped to ₹{sa_min:,}–₹{sa_max:,}).")
                df[sa_col] = pd.to_numeric(df[sa_col], errors='coerce').fillna(sum_assured)
                df[sa_col] = df[sa_col].clip(lower=sa_min, upper=sa_max)
                sa_series = df[sa_col]
            else:
                sa_series = pd.Series([sum_assured] * len(df), index=df.index)

            premium_main_list = []
            premium_co_list = []
            total_list = []
            statuses = []

            for idx, row in df.iterrows():
                row_status = "✅"
                p_main = None
                p_co = None
                row_sa = float(sa_series.loc[idx])
                try:
                    r_age = int(row[main_age_col])
                    r_tenure = int(row[main_tenure_col])
                    rate_main = get_rate(df_rates, tenure_map, r_age, r_tenure)
                    p_main = round(rate_main * (row_sa / 100000), 2)
                except Exception as e:
                    row_status = f"❌ Main Borrower: {e}"

                try:
                    r_age = int(row[co_age_col])
                    r_tenure = int(row[main_tenure_col])  # Co Borrower uses Main Borrower's tenure
                    rate_co = get_rate(df_rates, tenure_map, r_age, r_tenure)
                    p_co = round(rate_co * (row_sa / 100000), 2)
                except Exception as e:
                    row_status = (row_status + f" | Co Borrower: {e}") if row_status != "✅" else f"❌ Co Borrower: {e}"

                premium_main_list.append(p_main)
                premium_co_list.append(p_co)
                total_list.append(round(p_main + p_co, 2) if (p_main is not None and p_co is not None) else None)
                statuses.append(row_status)

            df["Main Borrower Premium"] = premium_main_list
            df["Co Borrower Premium"] = premium_co_list
            df["Total Premium"] = total_list
            df["Status"] = statuses

            core_cols = [
                main_name_col, main_age_col, main_tenure_col, "Main Borrower Premium",
                co_name_col, co_age_col, "Co Borrower Premium",
                "Total Premium"
            ]
            extra_cols = [c for c in df.columns if c not in core_cols]
            df_display = df[core_cols + extra_cols]

            grand_total = pd.to_numeric(pd.Series(total_list), errors='coerce').sum()
            st.metric("💰 Grand Total Premium", f"₹ {grand_total:,.2f}")

            st.subheader("Rate Lookup Output")
            st.dataframe(df_display, use_container_width=True)

            # ---- Build output with TOTAL PREMIUM row at the bottom ----
            total_row = {c: "" for c in df_display.columns}
            total_row[main_name_col] = "TOTAL PREMIUM"
            total_row["Total Premium"] = round(grand_total, 2)
            df_out = pd.concat([df_display, pd.DataFrame([total_row])], ignore_index=True)

            output_file = "Rate_Output.xlsx"

            # Write data starting from row 3 (1-indexed), leaving rows 1-2 for the 2-row header
            df_out.to_excel(output_file, index=False, header=False, startrow=2, sheet_name="Sheet1")

            wb = load_workbook(output_file)
            ws = wb["Sheet1"]

            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")

            n_extra = len(extra_cols)
            # Column positions (1-indexed): Main group = cols 1-4 (Name, Age, Tenure, Premium),
            # Co group = cols 5-7 (Name, Age, Premium — no Tenure since it's shared with Main),
            # Total Premium = col 8, extras start at col 9
            main_start, main_end = 1, 4
            co_start, co_end = 5, 7
            total_col = 8
            extra_start = 9

            # Row 2: sub-headers (actual field names)
            row2_labels = ["Name", "Age", "Tenure", "Premium", "Name", "Age", "Premium", "Total Premium"] + extra_cols
            for idx, label in enumerate(row2_labels, start=1):
                cell = ws.cell(row=2, column=idx, value=label)
                cell.font = bold
                cell.alignment = center

            # Row 1: group headers (merged)
            ws.merge_cells(start_row=1, start_column=main_start, end_row=1, end_column=main_end)
            ws.cell(row=1, column=main_start, value="MAIN BORROWER").font = bold
            ws.cell(row=1, column=main_start).alignment = center

            ws.merge_cells(start_row=1, start_column=co_start, end_row=1, end_column=co_end)
            ws.cell(row=1, column=co_start, value="CO BORROWER").font = bold
            ws.cell(row=1, column=co_start).alignment = center

            ws.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)
            ws.cell(row=1, column=total_col, value="TOTAL PREMIUM").font = bold
            ws.cell(row=1, column=total_col).alignment = center
            # remove duplicate row2 label under merged Total Premium cell
            ws.cell(row=2, column=total_col, value=None)

            if n_extra:
                ws.merge_cells(start_row=1, start_column=extra_start, end_row=1, end_column=extra_start + n_extra - 1)
                ws.cell(row=1, column=extra_start, value="ADDITIONAL INFO").font = bold
                ws.cell(row=1, column=extra_start).alignment = center

            # Auto width
            for col_idx in range(1, total_col + n_extra + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 18

            wb.save(output_file)

            with open(output_file, "rb") as file:
                st.download_button(
                    label="⬇ Download Output Excel",
                    data=file,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"Error: {e}")
